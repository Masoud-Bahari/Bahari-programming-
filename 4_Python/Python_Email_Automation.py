
#^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
import os
import sys
from openpyxl.styles import PatternFill
from numpy.core.defchararray import lower
from datetime import date
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formatdate
from email.mime.multipart import MIMEMultipart
from email import encoders
import os
from email.mime.application import MIMEApplication
from openpyxl.styles import PatternFill
import smtplib, ssl
import openpyxl as xl
from Email_Functions import Repeat_Email
from Email_Functions import Mass_Email
from Email_Functions import FollowUp
import Email_Functions
import arrow
#--------------------------------------------------------------------
TheExcelFile= xl.load_workbook(ExlName)
RepLis= xl.load_workbook(SuspListName)
#--------------------------------------------------------------------
#Avoiding repeatition
# Repeat_Email(RepLis, TheExcelFile, ExlName)
#--------------------------------------------------------------------
#Sending Email
Mass_Email(TheExcelFile, username, password, PdFName, PdFaddress,  ExlName)
#sheet1 = TheExcelFile['F1']
#--------------------------------------------------------------------
#Following up Email
#FollowUp(TheExcelFile, username, password, PdFName, PdFaddress, ExlName, ApplicantName)
#--------------------------------------------------------------------
# sheet1 = TheExcelFile['F1']
# print(sheet1['A{}'.format(68)].value)

#%%%%%%Functions
username='****@gmail.com' # str(input('Your sername:'))
password='@Fkl242019'#str(input('your password:'))
NameApp='Masoud Bahari'
ExlName='ShabnamList.xlsx'
PdFName='Masoud Bahari.pdf'
PdFName2="Masoud Bahari.pdf"
#--------------------------------------------------------------------
import smtplib, ssl
import openpyxl as xl
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formatdate
from email.mime.multipart import MIMEMultipart
from email import encoders
import os
from email.mime.application import MIMEApplication
from openpyxl.styles import PatternFill
from datetime import date
from datetime import datetime
import arrow
#--------------------------------------------------------------------
#--------------------------------------------------------------------
fcolor=PatternFill(patternType='solid', fgColor='03FF46')
fcolor1=PatternFill(patternType='solid', fgColor='FF9B03')
fcolor2=PatternFill(patternType='solid', fgColor='90FFFC')
today = date.today()
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
DATE1 = today.strftime("%b-%d-%Y")
#DATE2 = today.strftime("%d/%m/%Y")
TODAYY = arrow.get(today)
#--------------------------------------------------------------------
From = username
wb= xl.load_workbook(ExlName)
sheet1= wb['F1']
server= smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()#starting secure connection
server.login(username, password)
for i in range(2, len(sheet1['A'])+1):
 if sheet1['F{}'.format(i)].value != None:
  SentDate=arrow.get(sheet1['F{}'.format(i)].value)
  LengthDate=(TODAYY - SentDate).days
  if  LengthDate >= 6 and sheet1['G{}'.format(i)].value=='Waiting for report' and sheet1['G{}'.format(i)].value!=None:
       sheet1['G{}'.format(i)].fill = fcolor2
       sheet1['G{}'.format(i)].value='Followed Up, Date: '+ ' ('+DATE1+'), Time: ' + current_time
       msg=MIMEMultipart()
       msg['From']=username
       msg['To']=sheet1['A{}'.format(i)].value
       msg['Subject']='Following up my application: '+ NameApp
       text= ''' Dear professor {},\n 
   It has been a while since my query about PhD studies under your supervision and I have not heard back from you yet.
   Owing to my deep interest in your works, I would be very thankful if you could inform me about any possibilities to work under your supervision in your respected group.\n
   Cheers,
   Masoud
   Residential address: Via Francesco Ierace 3, 00133, Rome, Italy.'''.format(sheet1['B{}'.format(i)].value)

       part = MIMEBase('application', "octet-stream")
       part.set_payload(open(PdFName, "rb").read())
       encoders.encode_base64(part)
       part.add_header('Content-Disposition', 'attachment; filename={}'.format(PdFName2))
       msg.attach(part)
       msg.attach(MIMEText(text,'plain'))
       message=msg.as_string()
       server.sendmail(username, sheet1['A{}'.format(i)].value, message)
       print('Mail sent to', sheet1['A{}'.format(i)].value)
       wb.save(ExlName)
    #del msg['To']
server.quit()
print('DONE!!')
os.system("start EXCEL.EXE {}".format(ExlName))
#%%Functions
# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
fcolor = PatternFill(patternType='solid', fgColor='03FF46')
fcolor1 = PatternFill(patternType='solid', fgColor='FF9B03')
fcolor2 = PatternFill(patternType='solid', fgColor='90FFFC')
RedAlert = PatternFill(patternType='solid', fgColor='FF0000')
today = date.today()
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
DATE1 = today.strftime("%b-%d-%Y")
# DATE2 = today.strftime("%d/%m/%Y")
# DATE2 = today.strftime("%d/%m/%Y")
TODAYY = arrow.get(today)


# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%

def Repeat_Email(Suspected_List, TheExcelFile, ExlName):
    """
    Repeat_Email(Suspected_List,Email_List,ExlName),
    """
    Suspected = Suspected_List['F1']
    EmailMain = TheExcelFile['F1']
    # print(len(Suspected['A']))
    for i in range(1, len(EmailMain['A']) + 1):
        for j in range(1, len(Suspected['A']) + 1):
            if EmailMain['A{}'.format(i)].value != None and Suspected['A{}'.format(j)].value != None:
                St1 = str(lower(EmailMain['A{}'.format(i)].value))
                St2 = str(lower(Suspected['A{}'.format(j)].value))
                if EmailMain['E{}'.format(i)].value == None and St1 == St2 and Suspected['A{}'.format(j)].value != None:
                    EmailMain['A{}'.format(i)].value = 'Repeated ' + EmailMain['A{}'.format(i)].value
                    EmailMain['A{}'.format(i)].fill = RedAlert
                    print('Email List index=', i, 'Suspect List index=', j)
                    TheExcelFile.save(ExlName)
    # os.system("start EXCEL.EXE {}".format(ExlName))


# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%

def Mass_Email(TheExcelFile, username, password, PdFName, PdFaddress, ExlName):
    """
        Mass_Email(TheExcelFile,username,password,PdFName,ExlName):
        """

    import smtplib as smtp
    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)  # smtplib.SMTP('smtp.gmail.com', 587)
    #        server.starttls()  # starting secure connection
    server.login(username, password)
    sheet1 = TheExcelFile['F1']
    maintext = sheet1['H2'].value
    for i in range(1, len(sheet1['A']) + 1):
        if sheet1['E{}'.format(i)].value is None and sheet1['A{}'.format(i)].value is not None and sheet1['A{}'.format(i)].value.startswith("Repeated") == False:
            sheet1['E{}'.format(i)].fill = fcolor
            sheet1['E{}'.format(i)].value = 'Sent, Date: ' + '(' + DATE1 + '), Time: ' + current_time
            sheet1['F{}'.format(i)].value = today
            sheet1['G{}'.format(i)].value = 'Waiting for report'
            sheet1['G{}'.format(i)].fill = fcolor1
            proff_names = sheet1['B{}'.format(i)].value
            msg = MIMEMultipart()
            msg['From'] = username
            msg['To'] = sheet1['A{}'.format(i)].value.encode('ascii', 'ignore').decode('ascii')
            msg['Subject'] = sheet1['C{}'.format(i)].value
            text = f''' Dear professor {proff_names},\n\n {maintext}'''
            part = MIMEBase('application', "octet-stream")
            part.set_payload(open(PdFaddress, "rb").read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename={}'.format(PdFName))
            msg.attach(part)
            msg.attach(MIMEText(text, 'plain'))
            message = msg.as_string()
            server.sendmail(username, str(sheet1['A{}'.format(i)].value).encode('ascii', 'ignore').decode('ascii'),
                            message)
            print('Mail sent to', str(sheet1['A{}'.format(i)].value).encode('ascii', 'ignore').decode('ascii'))
            TheExcelFile.save(ExlName)
            # del msg['To']
    server.quit()
    print('Email process is DONE!!')
    TheExcelFile.save(ExlName)
    os.system("start EXCEL.EXE {}".format(ExlName))
# %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
def FollowUp(TheExcelFile, username, password, PdFName, PdFaddress, ExlName, ApplicantName):
    """
    FollowUp(TheExcelFile, username, password, PdFName, PdFaddress, ExlName, ApplicantName):
    """
    sheet1 = TheExcelFile['F1']
    maintext = sheet1['H3'].value
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()  # starting secure connection
    server.login(username, password)
    for i in range(2, len(sheet1['A']) + 1):
        if sheet1['F{}'.format(i)].value != None:
            SentDate = arrow.get(sheet1['F{}'.format(i)].value)
            LengthDate = (TODAYY - SentDate).days
            if LengthDate >= 6 and sheet1['G{}'.format(i)].value == 'Waiting for report' and sheet1[
                'G{}'.format(i)].value != None:
                sheet1['G{}'.format(i)].fill = fcolor2
                sheet1['G{}'.format(i)].value = 'Followed Up, Date: ' + ' (' + DATE1 + '), Time: ' + current_time
                proff_names = sheet1['B{}'.format(i)].value
                msg = MIMEMultipart()
                msg['From'] = username
                msg['To'] = str(sheet1['A{}'.format(i)].value).encode('ascii', 'ignore').decode('ascii')
                msg['Subject'] = 'Following up my application: ' + ApplicantName
                text = f''' Dear professor {proff_names},\n\n {maintext}'''
                part = MIMEBase('application', "octet-stream")
                part.set_payload(open(PdFaddress, "rb").read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment; filename={}'.format(PdFName))
                msg.attach(part)
                msg.attach(MIMEText(text, 'plain'))
                message = msg.as_string()
                server.sendmail(username, str(sheet1['A{}'.format(i)].value).encode('ascii', 'ignore').decode('ascii'),
                                message)
                print('Mail sent to', str(sheet1['A{}'.format(i)].value).encode('ascii', 'ignore').decode('ascii'))
                TheExcelFile.save(ExlName)
    server.quit()
    print('Following Up is DONE!!')
    os.system("start EXCEL.EXE {}".format(ExlName))


